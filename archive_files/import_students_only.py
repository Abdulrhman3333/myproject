"""
Import students from filled students_template.xlsx.
Run: python import_students_only.py
"""

import os
import sys
import uuid
from datetime import datetime

# Fix encoding on Windows for Arabic text
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

import django


def setup_django():
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "myproject.settings")
    django.setup()


def parse_date(value):
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or any(ch in text for ch in ["?", "#"]):
        return None

    formats = ["%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_phone(value):
    if value is None:
        return ""
    return str(value).strip()


def is_valid_identity(value):
    text = normalize_text(value)
    # Accept any non-empty identity number
    return len(text) > 0


def read_xlsx_rows(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required. Install it with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(item).strip() if item is not None else "" for item in rows[0]]
    result = []
    for values in rows[1:]:
        row_data = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = values[index] if index < len(values) else ""
            row_data[header] = "" if value is None else value
        result.append(row_data)
    return result


def import_students(base_dir, errors):
    from django.contrib.auth.models import User
    from quran_center.models import Student, LAST_TESTED_PART_CHOICES

    students_xlsx = os.path.join(base_dir, "students_template.xlsx")
    if not os.path.exists(students_xlsx):
        errors.append(f"Students file not found: {students_xlsx}")
        return {"created": 0, "updated": 0, "skipped": 0}

    grade_choices = {choice[0] for choice in Student.GRADE_CHOICES}
    status_choices = {choice[0] for choice in Student.STATUS_CHOICES}
    last_part_choices = {choice[0] for choice in LAST_TESTED_PART_CHOICES}
    default_status = Student._meta.get_field("status").get_default()

    created = 0
    updated = 0
    skipped = 0

    rows = read_xlsx_rows(students_xlsx)
    for index, row in enumerate(rows, start=2):
        full_name = normalize_text(row.get("full_name"))
        identity_number = normalize_text(row.get("identity_number"))
        parent_identity = normalize_text(row.get("parent_identity"))
        parent_phone = normalize_phone(row.get("parent_phone"))
        grade = normalize_text(row.get("grade"))
        birth_date_text = normalize_text(row.get("birth_date"))
        neighborhood = normalize_text(row.get("neighborhood"))

        # Skip only if BOTH full_name and identity_number are missing
        if not full_name and not identity_number:
            skipped += 1
            continue

        # If no identity_number, generate a unique one
        if not identity_number:
            identity_number = f"GEN_{uuid.uuid4().hex[:12]}"
        # If identity_number exists but invalid format, generate a unique one
        elif not is_valid_identity(identity_number):
            identity_number = f"GEN_{uuid.uuid4().hex[:12]}"

        # Validate grade if provided
        if grade and grade not in grade_choices:
            skipped += 1
            continue

        # Parse birth_date if provided
        birth_date = None
        if birth_date_text:
            birth_date = parse_date(birth_date_text)

        # Validate last_tested_part if provided
        last_tested_part = normalize_text(row.get("last_tested_part")) or "0"
        if last_tested_part not in last_part_choices:
            last_tested_part = "0"

        # Validate status if provided
        status = normalize_text(row.get("status")) or default_status
        if status not in status_choices:
            status = default_status

        # Get teacher if provided
        teacher_username = normalize_text(row.get("teacher_username"))
        teacher = None
        if teacher_username:
            teacher = User.objects.filter(username=teacher_username).first()

        defaults = {
            "full_name": full_name or f"Student {identity_number}",
            "student_phone": normalize_phone(row.get("student_phone")),
            "parent_phone": parent_phone,
            "parent_identity": parent_identity,
            "grade": grade or "",
            "birth_date": birth_date,
            "last_tested_part": last_tested_part,
            "previous_center": normalize_text(row.get("previous_center")),
            "neighborhood": neighborhood,
            "status": status,
            "teacher": teacher,
        }

        student, was_created = Student.objects.get_or_create(
            identity_number=identity_number,
            defaults=defaults,
        )

        if was_created:
            created += 1
            print(f"[+] Created: {defaults['full_name']}")
        else:
            updated += 1
            for key, value in defaults.items():
                setattr(student, key, value)
            student.save()
            print(f"[+] Updated: {defaults['full_name']}")

    return {"created": created, "updated": updated, "skipped": skipped}


def main():
    base_dir = os.path.join(os.path.dirname(__file__), "import_templates")
    errors = []

    print("Importing students from students_template.xlsx...\n")
    result = import_students(base_dir, errors)

    print(f"\n{'='*50}")
    print(f"Students: created={result['created']}, updated={result['updated']}, skipped={result['skipped']}")
    print(f"Total processed: {result['created'] + result['updated'] + result['skipped']}")
    
    total = result['created'] + result['updated']
    if total > 0:
        print(f"\n[OK] Successfully imported {total} student(s)!")
    else:
        print("\n[ERROR] No students were imported.")


if __name__ == "__main__":
    setup_django()
    main()
